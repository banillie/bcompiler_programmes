from bcompiler.utils import project_data_from_master
from openpyxl import load_workbook, Workbook

def get_emails(master_dict, email_keys, email_addresses):
    output_list = []

    for project_name in master_dict.keys():
        for key in email_keys:
            output_list.append(master_dict[project_name][key])

    output_list = list(set(output_list))

    for email in output_list:
        if email == None:
            output_list.remove(email)

    final_list = []

    for email in output_list:
        for address in email_addresses:
            if address in email:
                final_list.append(email)

    return final_list

def parse_into_excle(email_list):
    wb = Workbook()
    ws = wb.active

    for i, email in enumerate(email_list):
        ws.cell(row=i+1, column=1).value = email

    return wb

master = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx')
email_keys = ['Working Contact Email', 'SRO Email', 'PD Email' ]
addresses = ['dft', 'dvsa', 'highwaysengland']

test = get_emails(master, email_keys, addresses)

parse = parse_into_excle(test)
parse.save('C:\\Users\\Standalone\\Will\\email_list.xlsx')