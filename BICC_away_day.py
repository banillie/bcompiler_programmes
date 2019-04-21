'''I used this programme for the BICC away day 2018'''


from bcompiler.utils import project_data_from_master
from openpyxl import load_workbook, Workbook

def pie_chart_info(m_one, m_two, m_three, m_four, dca_of_interest):
    master_list = [m_one, m_two, m_three, m_four]

    wb = Workbook()
    ws = wb.active

    row = 1
    for master in master_list:
        dca_list = []
        for name in master.keys():
            dca = master[name][dca_of_interest]
            if dca != None:
                dca_list.append(dca)

        values = sorted(list(set(dca_list)))
        #print(values)
        #print(master[name]['Reporting period (GMPP - Snapshot Date)'])

        wlc_per_rating = calculating_rating_value(master, values, dca_of_interest, remove)

        row += 1
        ws.cell(row=row + 1, column=1).value = master[name]['Reporting period (GMPP - Snapshot Date)']
        row += 1
        for i in range(len(values)):
            ws.cell(row=row + 1, column=1).value = values[i]
            ws.cell(row=row + 1, column=2).value = dca_list.count(values[i])
            ws.cell(row=row + 1, column=3).value = wlc_per_rating[i]
            row += 1

    return wb

def calculating_rating_value(master, values_list, dca_of_interest, avoid_double_counting):
    overall_cost_list = []
    for value in values_list:
        cost_list = []
        for name in master.keys():

            if name in avoid_double_counting:
                pass
            else:
                if master[name][dca_of_interest] == value:
                    cost = master[name]['Total Forecast']
                    if cost != None:
                        cost_list.append(int(cost))

            totals = sum(cost_list)

        overall_cost_list.append(totals)

    return overall_cost_list


def number_of_projects(m_one, m_two, m_three, m_four):
    four = list(m_four.keys())
    three = list(m_three.keys())
    two = list(m_two.keys())
    one = list (m_one.keys())

    left_in_q3 = [x for x in four if x not in three]
    joined_in_q3 = [x for x in three if x not in four]
    left_in_q4 = [x for x in three if x not in two]
    joined_in_q4 = [x for x in two if x not in three]
    left_in_q1 = [x for x in two if x not in one]
    joined_in_q1 = [x for x in one if x not in two]
    left_in_q2 = [x for x in one if x not in zero]
    joined_in_q2 = [x for x in zero if x not in one]


    print('Q3 left' + str(left_in_q3))
    print('Q4 joined' + str(joined_in_q3))
    print('Q4 left ' + str(left_in_q4))
    print('Q1 joined ' + str(joined_in_q4))
    print('Q1 left ' + str(left_in_q1))
    print('Q2 joined ' + str(joined_in_q1))


zero = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx')
one = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_1_2018.xlsx')
two = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_4_2017.xlsx')
three = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2017.xlsx')

remove = ['Cross Country Rail Franchise Competition', 'East Midlands Franchise', 'South Eastern Rail Franchise Competition',
          'West Coast Partnership Franchise', 'HS2 Phase 2b', 'HS2 Phase1', 'HS2 Phase2a']

#output = pie_chart_info(three, two, one, zero, 'Overall Resource DCA - Now')
#output.save('C:\\Users\\Standalone\\Will\\BICC_away_resourcing_dca.xlsx')

number_of_projects(zero, one, two, three)