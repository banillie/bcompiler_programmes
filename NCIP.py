'''I used this programme for comparing national construction and infrastructure pipeline and gmpp data'''

from bcompiler.utils import project_data_from_master
from openpyxl import load_workbook, Workbook
#import datetime
#from openpyxl.utils import column_index_from_string
#from collections import OrderedDict

def placing_excel(wb, portfolio_master):
    ws = wb.active

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=6).value
        if project_name in portfolio_master:
            ws.cell(row=row_num, column=8).value = portfolio_master[project_name]['Delivery Structure']
            ws.cell(row=row_num, column=11).value = portfolio_master[project_name]['Brief project description (GMPP - brief descripton)']
            ws.cell(row=row_num, column=16).value = portfolio_master[project_name]['Source of Finance']
            ws.cell(row=row_num, column=19).value = portfolio_master[project_name]['Project MM19 Original Baseline'] #start of construction
            ws.cell(row=row_num, column=21).value = portfolio_master[project_name]['Project MM20 Original Baseline'] # start of operation
            ws.cell(row=row_num, column=23).value = portfolio_master[project_name]['Total Budget/BL']
            total = int(portfolio_master[project_name]['17-18 RDEL BL Total']) + int(portfolio_master[project_name]['17-18 CDEL BL Total']) + int(portfolio_master[project_name]['17-18 BL Non-Gov'])
            ws.cell(row=row_num, column=27).value = total
            total = portfolio_master[project_name]['18-19 RDEL BL Total'] + portfolio_master[project_name]['18-19 CDEL BL Total'] + portfolio_master[project_name]['18-19 BL Non-Gov']
            ws.cell(row=row_num, column=29).value = total
            total = portfolio_master[project_name]['19-20 RDEL BL Total'] + portfolio_master[project_name]['19-20 CDEL BL Total'] + portfolio_master[project_name]['19-20 BL Non-Gov']
            ws.cell(row=row_num, column=31).value = total
            total = portfolio_master[project_name]['20-21 RDEL BL Total'] + portfolio_master[project_name]['20-21 CDEL BL Total'] + portfolio_master[project_name]['20-21 BL Non-Gov']
            ws.cell(row=row_num, column=33).value = total
            total = portfolio_master[project_name]['21-22 RDEL BL Total'] + portfolio_master[project_name][
                '21-22 CDEL BL Total'] + portfolio_master[project_name]['21-22 BL Non-Gov']
            ws.cell(row=row_num, column=35).value = total
            ws.cell(row=row_num, column=37).value = portfolio_master[project_name]['Real or Nominal - Baseline']

        else:
            print('could find ' + str(project_name))

    return wb

def projects_missing(wb, master):
    master_list = master.keys()

    worksheet_list = []
    ws = wb.active
    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=6).value
        worksheet_list.append(project_name)

    missing = [x for x in master_list if x not in worksheet_list]

    print(missing)
    print(worksheet_list)

master = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018_for_ncip_comparison_to_delete.xlsx')
ncip_wb = load_workbook('C:\\Users\\Standalone\\Will\\Ncip_comparison.xlsx')

populated_ncip_wb = placing_excel(ncip_wb, master)
populated_ncip_wb.save('C:\\Users\\Standalone\\Will\\ncip_comparison_ver3.xlsx')

projects_missing(ncip_wb, master)