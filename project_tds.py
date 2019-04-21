'''This programme produces project specific time delta analysis outputs that require some further manual edits
so they are understandable. The purpose of this analysis is to understand how far standard project milestones
are moving away from project baselines.

This programme is working well. However in some instances it doesn't produce the desired output. This is most likely
due to the data being reported by projects, needing to be checked and amended in the master.

notes on now programme works:
1) master data is converted into python dictionary format
2) these dictionaries are put into a list
3) when sorting data there are two key value pairs at the beginning. i) Milestone keys : milestone names ii) milestone
actual - forecast : milestone dates. the two key value pairs are also out into lists.
4) the above dictionaries, together with a list of project names (which is taken from latest master dictionary)
are passed in the milestone_extraction function.
5) the milestone_extraction function takes this data. It firstly uses the
    a) bc_ref_stages function to return a dictionary structured in the following way project name[('latest quarter info'
    , 'latest bc'), ('last quarter info', 'last bc'),('last baseline quarter info', 'last baseline bc'),
    ('oldest quarter info', 'oldest bc')] depending on the amount information available in the data. Only the first
    three key values are returned, to ensure consistency (which is helpful later).
The milestone_extraction function itself returns a dictionary containing the project_name[('latest quarter',
('project start', datetime), ('latest quarter', ('SOBC', datetime))... continued until ('latest quarter', ('end project'
, datetime))... then ('last quarter', ('project start', datetime)) etc... then ('baseline quarter', ('project start',
datetime)) etc...
6) This dictionary is then passed into the cal_td (calculate_timedelta) function. This function calculates the time
delta (td) / amount of time between milestone stages. e.g. the amount of time between sobc and obc. It removes/doesn't
calculate 'start of project' tds (as they don't exist i.e. no earlier milestone to use to calculate td). The
cal_td function returns a dictionary in same format as above but td information is stored in the datetime section. e.g.
project_name[('quarter info', ('stage', datetime.timedelta) etc...]
7) the milestone and td dictionaries are passed into the run function which opens a excle workbook and parses data into
the workbook via..
8) the placing_date_in_excel_single function. this function uses the following functions:
    a) sort_td. Arranges date so milestones are in order sobc, sobc, sobc, obc, obc, obc... etc. The data is then placed
    in the excel wb in this order - for the swimlane chart.
    b) bc_list. returns list of business cases in the correct order i.e. [SOBC, OBC, FBC, SoC, SoP, End]
9) within the run programme the latest approved bc stage of the project is calculated.
10) outputs at 8 and 9 are parsed into the build_chart_single function, which creates the chart and returns the saved
workbook.

todo amend code so that the charts build place latest, last quarter, and baseline milestone data '''
from typing import List, Any, Tuple

from bcompiler.utils import project_data_from_master
from openpyxl import Workbook
import datetime
from openpyxl.chart import ScatterChart, Reference, Series

'''The milestone_extraction function itself returns a dictionary containing the project_name[('latest quarter',
('project start', datetime), ('latest quarter', ('SOBC', datetime))... continued until ('latest quarter', ('end project'
, datetime))... then ('last quarter', ('project start', datetime)) etc... then ('baseline quarter', ('project start',
datetime)) etc...'''

def milestone_extraction(projects, master_data_all, milestone_keys, milestone_dates):
    global project, quarter_master
    output_dict = {}

    '''firstly business cases of interest are filtered out by bc_ref_stage function'''
    list_of_bc_stages = bc_ref_stages(projects, master_data_all)

    for project in projects:
        print(project)
        td_list = []
        bc_interest = list_of_bc_stages[project]
        for i in range(0, len(bc_interest)):
            for master in master_data_all:
                try:
                    date = master[project]['Reporting period (GMPP - Snapshot Date)']

                    if date == bc_interest[i][0]:
                        for x in range(len(milestone_keys)):
                            key_1 = milestone_keys[x]
                            milestone_name = master[project][key_1]
                            key_2 = milestone_dates[x]
                            milestone_date = master[project][key_2]
                            td_info = (milestone_name, milestone_date)
                            td_list.append((date, td_info))
                except KeyError:
                    pass

        output_dict[project] = td_list

    return output_dict

''' function to returns a dictionary structured in the following way project name[('latest quarter info', 'latest bc'), 
('last quarter info', 'last bc'), ('last baseline quarter info', 'last baseline bc'), ('oldest quarter info', 
'oldest bc')] depending on the amount information available in the data. Only the first three key values are returned, 
to ensure consistency (which is helpful later).'''
def bc_ref_stages(project_name_list, master_list):

    output_dict = {}

    for name in project_name_list:
        all_list = []      # format [('quarter info': 'bc')] across all masters including project
        bl_list = []        # format ['bc', 'bc'] across all masters. bl_list_2 removes duplicates
        ref_list = []       # format as for all list but only contains the three tuples of interest
        for master in master_list:
            try:
                bc_stage = master[name]['BICC approval point']
                quarter = master[name]['Reporting period (GMPP - Snapshot Date)']
                tuple = (quarter, bc_stage)
                all_list.append(tuple)
            except KeyError:
                pass

        for i in range(0, len(all_list)):
            bl_list.append(all_list[i][1])


        '''below lines of text from stackoverflow. Question, remove duplicates in python list while 
        preserving order'''
        seen = set()
        seen_add = seen.add
        bl_list_2 = [x for x in bl_list if not (x in seen or seen_add(x))]

        ref_list.insert(0, all_list[0])     # puts the latest info into the list first

        try:
            ref_list.insert(1, all_list[1])    # puts that last info into the list
        except IndexError:
            ref_list.insert(1, all_list[0])

        if len(bl_list_2) == 1:                     # puts oldest info into list (if no baseline)
            ref_list.insert(2, all_list[-1])
        else:
            for i in range(0, len(all_list)):      # puts in baseline
                if all_list[i][1] == bl_list[0]:
                    ref_list.insert(2, all_list[i])

        '''that is a hack here i.e. returning only first three in ref_list. There's a bug which I don't fully 
        understand, but this solution is hopefully good enough for now'''
        output_dict[name] = ref_list[0:3]

    return output_dict

''' Arranges date so milestones are in order sobc, sobc, sobc, obc, obc, obc... etc. The data is then placed
    in the excel wb in this order - for the swimlane chart. '''
def sort_td(milestone_list):
    list_of_bc = bc_list(milestone_list)
    #print(milestone_list)
    #print(list_of_bc)

    output_list = []
    '''couldn't find a one loop solution to below in time available'''
    for i in range(0, 3):
        '''the hard code for loop is essentially designed as (i, 18+i, 6)'''
        for x in range(i, (len(milestone_list) + i), len(list_of_bc)):
            output_list.append(milestone_list[x])
    for i in range(3, 6):
        for x in range(i, (len(milestone_list) + i), len(list_of_bc)):
            output_list.append(milestone_list[x])

    return output_list


'''function that compiles list of all business case stages being reported and puts then in the right order'''
def bc_list(milestone_list):
    output_list = []
    for bc in milestone_list:
        output_list.append(bc[1][0])
    output_list = sorted(set(output_list), key=output_list.index)

    return output_list

''' Fuction places data into an excle wb in the order required for the swimlane chart, plus all raw data so it can 
be seen in the wb.'''
def placing_date_in_excel_single(ws, project_name, milestone_data_dict, td_data_dict):

    milestone_master = milestone_data_dict[project_name]

    '''placing data into the worksheet, so it can be seen'''
    for i in range(len(milestone_master)):
        ws.cell(row=i + 1, column=1).value = milestone_master[i][0]
        ws.cell(row=i + 1, column=2).value = milestone_master[i][1][0]
        ws.cell(row=i + 1, column=3).value = milestone_master[i][1][1]

    '''placing time delta data into worksheet. firstly the above function is used to sort the data'''

    td_master = td_data_dict[project_name]
    sorted_td = sort_td(td_master)

    for i in range(len(sorted_td)):
        ws.cell(row=i + 1, column=5).value = sorted_td[i][0]
        ws.cell(row=i + 1, column=6).value = sorted_td[i][1][0]
        td = sorted_td[i][1][1]
        ws.cell(row=i + 1, column=7).value = td.days

    '''the below loop provides index numbers for the chart lanes'''
    list_of_bc = bc_list(td_master)

    a = int(len(sorted_td)/len(list_of_bc))

    start_row = 1
    for i in range(len(list_of_bc)):
        for x in range(a):
            ws.cell(row= start_row, column= 8).value = i+1
            start_row += 1

    return ws

'''This function calculates the time delta (td) / amount of time between milestone stages. e.g. the amount of time 
between sobc and obc. It removes/doesn't calculate 'start of project' tds (as they don't exist i.e. no earlier milestone
to use to calculate td). The cal_td function returns a dictionary in same format as above but td information is stored 
in the datetime section. e.g. project_name[('quarter info', ('stage', datetime.timedelta) etc...]'''
def cal_td(milestone_data_dict):
    output_dict = {}

    for name in milestone_data_dict.keys():
        project_master = milestone_data_dict[name]
        output_list = []
        for i in range(len(project_master)-1):
            '''exception rule required here to handle the final time delta calculation'''
            quarter_stamp = project_master[i+1][0]
            milestone_info = project_master[i+1][1][0]
            try:
                td = project_master[i+1][1][1] - project_master[i][1][1]
            except TypeError:
                td = datetime.timedelta(0)
            td_info = (milestone_info, td)
            output_list.append((quarter_stamp, td_info))
        '''removal of start of project time deltas. The first one is not calculated
        note hard coded '''
        for i in output_list:
            if 'Start of Project' in i[1]:
                output_list.remove(i)

        output_dict[name] = output_list

    return output_dict

'''function required for finding the min td for projects. This is used to help set the x-axis boundaires'''
def min_value(project_name, td_data_dict):
    td_master = td_data_dict[project_name]
    output_list = []
    for i in range(len(td_master)):
        output_list.append(td_master[i][1][1].days)

    return min(output_list)

'''function required for find the max td for projects. This is used to help set x-axis boundaries'''
def max_value(project_name, td_data_dict):
    td_master = td_data_dict[project_name]
    output_list = []
    for i in range(len(td_master)):
        output_list.append(td_master[i][1][1].days)

    return max(output_list)

'''Function creates the swimlane chart and then returns the wb. 

Further work could be done here to tidy up the output'''
def build_chart_single(ws, project_name, approval_point, td_data_dict):
    chart = ScatterChart()
    chart.title = str(project_name) + ' last approved business case: ' + str(approval_point)
    chart.style = 18
    chart.x_axis.title = 'Time delta for each business case (year intervals)'
    #chart.y_axis.title = 'Milestones'
    chart.auto_axis = False
    '''this code is necessary to calculate min chart value if its greater than zero'''
    x_axis_min = min_value(project_name, td_data_dict)
    if x_axis_min >= 0:
        chart.x_axis.scaling.min = 0
    elif x_axis_min < 0:
        anchor =  x_axis_min % 365
        chart.x_axis.scaling.min = x_axis_min - anchor
    chart.x_axis.scaling.max = max_value(project_name, td_data_dict)  # max number (of days) in the x axis. calculated by max_value function
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 7  # hard coded for now - although minor issue as number of bc time deltas static
    chart.height = 9  # default is 7.5
    chart.width = 21  # default is 15

    '''changes units on x and y axis'''
    chart.x_axis.majorUnit = 365  # hard coded for now - minor issue as td will normally be in year intervals
    # chart.y_axis.majorUnit = 1.0   testing to see if required

    '''reverses y axis'''
    #chart.x_axis.scaling.orientation = "minMax"
    #chart.y_axis.scaling.orientation = "maxMin"

    '''makes the x axis cross at the max y value'''
    #chart.x_axis.crosses = 'max'

    '''removes lable on y axis'''
    chart.y_axis.delete = True

    #TOD: sort styling
    '''styling chart'''
    '''formating for titles'''
    #font = Font(typeface='Calibri')
    #size = 1200  # 12 point size
    #cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
    #pp = ParagraphProperties(defRPr=cp)
    #rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    #chart.x_axis.title.tx.rich.p[0].pPr = pp  # x_axis title

    #size_2 = 1400
    #cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
    #pp_2 = ParagraphProperties(defRPr=cp_2)
    #rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
    #chart.title.tx.rich.p[0].pPr = pp_2  # chart title

    '''the below assigns series information to the data that has been placed in the chart. 
    old values are placed first show that they show behind the current values'''

    for i in range(0,18,3):
        xvalues = Reference(ws, min_col=7, min_row=i+1, max_row=i+1)
        yvalues = Reference(ws, min_col=8, min_row=i+1, max_row=i+1)
        series = Series(values=yvalues, xvalues=xvalues, title="Latest quarter")
        chart.series.append(series)
        s1 = chart.series[i]
        s1.marker.symbol = "diamond"
        s1.marker.size = 10
        s1.marker.graphicalProperties.solidFill = "c9e243"  # Marker filling greenish
        s1.marker.graphicalProperties.line.solidFill = "c9e243"  # Marker outline greenish
        s1.graphicalProperties.line.noFill = True

        xvalues = Reference(ws, min_col=7, min_row=i + 2, max_row=i + 2)
        yvalues = Reference(ws, min_col=8, min_row=i + 2, max_row=i + 2)
        series = Series(values=yvalues, xvalues=xvalues, title="Last quarter")
        chart.series.append(series)
        s1 = chart.series[i+1]
        s1.marker.symbol = "diamond"
        s1.marker.size = 10
        s1.marker.graphicalProperties.solidFill = "ced0ff"  # Marker filling grey/blue
        s1.marker.graphicalProperties.line.solidFill = "ced0ff"  # Marker outline grey/blue
        s1.graphicalProperties.line.noFill = True

        xvalues = Reference(ws, min_col=7, min_row=i + 3, max_row=i + 3)
        yvalues = Reference(ws, min_col=8, min_row=i + 3, max_row=i + 3)
        series = Series(values=yvalues, xvalues=xvalues, title="Baseline")
        chart.series.append(series)
        s1 = chart.series[i + 2]
        s1.marker.symbol = "diamond"
        s1.marker.size = 10
        s1.marker.graphicalProperties.solidFill = "8187ff"  # Marker filling blue
        s1.marker.graphicalProperties.line.solidFill = "8187ff"  # Marker outline blue
        s1.graphicalProperties.line.noFill = True


    ws.add_chart(chart, "K2")

    return ws

'''function for running individual project td outputs'''
def run(project_name, milestone_data, td_data):
    wb = Workbook()
    ws = wb.active
    parser_1 = placing_date_in_excel_single(ws, project_name, milestone_data, td_data)
    approval_point = zero[project_name]['BICC approval point']

    build_chart_single(parser_1, project_name, approval_point, td_data)

    return wb

'''master files are loaded here'''
zero_2 = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2018.xlsx')
zero = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx')
one = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_1_2018.xlsx')
two = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_4_2017.xlsx')
three = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2017.xlsx')
four = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2017.xlsx')
five = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_1_2017.xlsx')
six = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_4_2016.xlsx')
last = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2016.xlsx')

'''master files put into a list'''
master_list_all = [zero_2, zero, one, two, three, four, five, six, last]
#master_list_two = [zero, last]

project_names = zero.keys()
#project_names = ['Network Rail Asset Disposal']

milestone_keys = ['Project MM18', 'Approval MM1', 'Approval MM3', 'Approval MM10', 'Project MM19', 'Project MM20', 'Project MM21']
milestone_dates = ['Project MM18 Forecast - Actual', 'Approval MM1 Forecast / Actual', 'Approval MM3 Forecast / Actual', 'Approval MM10 Forecast / Actual',
                   'Project MM19 Forecast - Actual', 'Project MM20 Forecast - Actual', 'Project MM21 Forecast - Actual']

milestone_data = milestone_extraction(project_names, master_list_all, milestone_keys, milestone_dates)

td_data = cal_td(milestone_data)

'''run individual project tds analysis'''
for project_name in milestone_data.keys():
    out = run(project_name, milestone_data, td_data)
    out.save('C:\\Users\\Standalone\\Will\\{}_td_analysis.xlsx'.format(project_name))