'''Programme for producing reference class forecastings analysis

todo go through code and tidy up
todo create a seperate programme for time delta analysis - first'''


from bcompiler.utils import project_data_from_master
from openpyxl import load_workbook, Workbook
import datetime
#from openpyxl.utils import column_index_from_string
from openpyxl.chart import ScatterChart, Reference, Series
#from openpyxl.chart.text import RichText
#from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties, Font, RegularTextRun


def milestone_extraction(projects, master_data_all, milestone_keys, milestone_dates):
    global project, quarter_master
    output_dict = {}

    list_of_bc_stages = bc_ref_stages(projects, master_data_all)
    print(list_of_bc_stages)

    for project in projects:
        print(project)
        td_list = []
        bc_interest = list_of_bc_stages[project]
        for i in range(2):     # just two as only want the latest of the last bc change/or oldest information
            for master in master_data_all:
                try:
                    date = master[project]['Reporting period (GMPP - Snapshot Date)']

                    if date == bc_interest[i][0]:
                        for x in range(len(milestone_keys)):
                            key_1 = milestone_keys[x]
                            milestone_name = master[project][key_1]
                            key_2 = milestone_dates[x]
                            milestone_date = master[project][key_2]
                            td_info = (milestone_name, milestone_date)
                            td_list.append((date, td_info))
                except KeyError:
                    pass

        output_dict[project] = td_list

    return output_dict

'''function for calculating current and last bc stage'''
def bc_ref_stages(project_name_list, master_list):
    '''first stage... get list of all bc stages reported by the project'''
    first_dict  = {}
    for project in project_name_list:
        lower_list = []
        for master in master_list:
            try:
                bc_stage = master[project]['BICC approval point']
                quarter = master[project]['Reporting period (GMPP - Snapshot Date)']
                tuple = (quarter, bc_stage)
                lower_list.append(tuple)
            except KeyError:
                pass
        first_dict[project] = lower_list

    '''second stage... sorting the list based on when bc stage changes'''

    second_dict = {}
    for name in first_dict.keys():
        project_list = first_dict[name]
        lower_list = []
        for i in range(1,(len(project_list))):    # take out the first/latest info as put in later
            try:
                if project_list[i][1] != project_list[i+1][1]:
                    lower_list.append(project_list[i])

            except IndexError:
                pass

        lower_list.insert(0, project_list[0])     # puts the first/lastest info into the list first

        if lower_list[-1] != project_list[-1]:      # checking to see that there is oldest info in there
            lower_list.append(project_list[-1])

        second_dict[name] =  lower_list

    return second_dict


'''function for calculating the time delta between milestones'''
def cal_td(milestone_data_dict):
    output_dict = {}

    for name in milestone_data_dict.keys():
        project_master = milestone_data_dict[name]
        output_list = []
        for i in range(len(project_master)-1):
            '''exception rule required here to handle the final time delta calculation'''
            quarter_stamp = project_master[i+1][0]
            milestone_info = project_master[i+1][1][0]
            try:
                td = project_master[i+1][1][1] - project_master[i][1][1]
            except TypeError:
                td = datetime.timedelta(0)
            td_info = (milestone_info, td)
            output_list.append((quarter_stamp, td_info))
        '''removal of the second start of project. The first one is not calculated'''
        output_list.remove(output_list[6])

        output_dict[name] = output_list

    return output_dict

'''function that compiles list of all business cases stage names being reported'''
def bc_list(td_data_dict):
    output_list = []
    data_list = td_data_dict
    for bc in data_list:
        output_list.append(bc[1][0])
    output_list = sorted(set(output_list), key=output_list.index)

    return output_list

'''started this and then stopped as the above function is ok for now. worth looking at this in more detail 
at some point though as might be handy of the programme is going to look at all milestones'''
#def bc_list_2(project_name, milestone_keys, milestone_data_dict):
#    milestone_master = milestone_data_dict[project_name]

'''function that returns all the different business case stages being reported by projects'''
def bc_stages(master_dictionary):
    output_list = []
    for name in master_dictionary.keys():
        master = master_dictionary[name]
        for key in master.keys():
            if key == 'BICC approval point':
                print('yes')
                output_list.append(master_dictionary[name][key])

    duplicates_removed = list(set(output_list))

    return duplicates_removed

'''function that compiles list of all quarters that data has been reported for'''
def quarter_list(project_name, td_data_dict):
    output_list = []
    data_list = td_data_dict[project_name]
    for quarter in data_list:
        output_list.append(quarter[0])
    sorted_list = sorted(set(output_list), key=output_list.index)

    return sorted_list

'''this function sorts the td_data_dictionaries in order of BC. So its is SOBCss first, then OBCs ect'''
def sort_td(td_data_dict):
    list_of_bc = bc_list(td_data_dict)
    #print(list_of_bc)
    td_master = td_data_dict
    #print(td_master)

    output_list = []
    for i in range(len(list_of_bc)):
        '''the below loop was designed as (i, 12+i, 6)
        but should be flexible to handle a larger td_master 
        that goes over more than one quarter'''
        for x in range(i, (len(td_master) + i), len(list_of_bc)):
            output_list.append(td_master[x])

    return output_list



def placing_date_in_excel_single(ws, project_name, milestone_data_dict, td_data_dict):

    milestone_master = milestone_data_dict[project_name]
    '''placing data into the worksheet, so it can be seen'''
    for i in range(len(milestone_master)):
        ws.cell(row=i + 1, column=1).value = milestone_master[i][0]
        ws.cell(row=i + 1, column=2).value = milestone_master[i][1][0]
        ws.cell(row=i + 1, column=3).value = milestone_master[i][1][1]

    '''placing time delta data into worksheet. 
    firstly the above function is used to sort the data'''

    td_master = td_data_dict[project_name]

    sorted_td = sort_td(td_master)

    for i in range(len(sorted_td)):
        ws.cell(row=i + 1, column=5).value = sorted_td[i][0]
        ws.cell(row=i + 1, column= 6).value = sorted_td[i][1][0]
        td = sorted_td[i][1][1]
        ws.cell(row=i + 1, column=7).value = td.days

    '''the below loop provides index numbers for the chart lanes'''
    list_of_bc = bc_list(td_master)

    a = int(len(sorted_td)/len(list_of_bc))

    start_row = 1
    for i in range(len(list_of_bc)):
        for x in range(a):
            ws.cell(row= start_row, column= 8).value = i+1
            start_row += 1

    return ws

'''function required for find the max td for projects. This is used to help set x-axis boundaries'''
def max_value(project_name, td_data_dict):
    td_master = td_data_dict[project_name]
    output_list = []
    for i in range(len(td_master)):
        output_list.append(td_master[i][1][1].days)

    return max(output_list)

'''function required for finding the min td for projects. This is used to help set the x-axis boundaires'''
def min_value(project_name, td_data_dict):
    td_master = td_data_dict[project_name]
    output_list = []
    for i in range(len(td_master)):
        output_list.append(td_master[i][1][1].days)

    return min(output_list)

'''function required for finding the min td for projects. This is used to help set the x-axis boundaires.
for aggregate chart'''
def min_value_all(td_data_dict):
    output_list = []

    for name in td_data_dict.keys():
        td_master = td_data_dict[name]
        for i in range(len(td_master)):
            output_list.append(td_master[i][1][1].days)

    return min(output_list)

'''function required for finding the max td for projects. This is used to help set the x-axis boundaires. 
for aggregate chart'''
def max_value_all(td_data_dict):
    output_list = []

    for name in td_data_dict.keys():
        td_master = td_data_dict[name]
        for i in range(len(td_master)):
            output_list.append(td_master[i][1][1].days)

    return max(output_list)

def build_chart_single(ws, project_name, approval_point, td_data_dict):
    chart = ScatterChart()
    chart.title = str(project_name) + ' last approved business case: ' + str(approval_point)
    chart.style = 18
    chart.x_axis.title = 'Time delta for each business case (year intervals)'
    #chart.y_axis.title = 'Milestones'
    chart.auto_axis = False
    '''this code is necessary to calculate min chart value if its greater than zero'''
    x_axis_min = min_value(project_name, td_data_dict)
    if x_axis_min >= 0:
        chart.x_axis.scaling.min = 0
    elif x_axis_min < 0:
        anchor =  x_axis_min % 365
        chart.x_axis.scaling.min = x_axis_min - anchor
    chart.x_axis.scaling.max = max_value(project_name, td_data_dict)  # max number (of days) in the x axis. calculated by max_value function
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 7  # hard coded for now - although minor issue as number of bc time deltas static
    chart.height = 9  # default is 7.5
    chart.width = 21  # default is 15

    '''changes units on x and y axis'''
    chart.x_axis.majorUnit = 365  # hard coded for now - minor issue as td will normally be in year intervals
    # chart.y_axis.majorUnit = 1.0   testing to see if required

    '''reverses y axis'''
    #chart.x_axis.scaling.orientation = "minMax"
    #chart.y_axis.scaling.orientation = "maxMin"

    '''makes the x axis cross at the max y value'''
    #chart.x_axis.crosses = 'max'

    '''removes lable on y axis'''
    chart.y_axis.delete = True

    #TOD: sort styling
    '''styling chart'''
    '''formating for titles'''
    #font = Font(typeface='Calibri')
    #size = 1200  # 12 point size
    #cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
    #pp = ParagraphProperties(defRPr=cp)
    #rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    #chart.x_axis.title.tx.rich.p[0].pPr = pp  # x_axis title

    #size_2 = 1400
    #cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
    #pp_2 = ParagraphProperties(defRPr=cp_2)
    #rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
    #chart.title.tx.rich.p[0].pPr = pp_2  # chart title

    '''the below assigns series information to the data that has been placed in the chart. 
    old values are placed first show that they show behind the current values'''

    for i in range(0,12,2):
        xvalues = Reference(ws, min_col=7, min_row=i+1, max_row=i+1)
        yvalues = Reference(ws, min_col=8, min_row=i+1, max_row=i+1)
        series = Series(values=yvalues, xvalues=xvalues, title="Current/Actual")
        chart.series.append(series)
        s1 = chart.series[i]
        s1.marker.symbol = "diamond"
        s1.marker.size = 10
        s1.marker.graphicalProperties.solidFill = "c9e243"  # Marker filling greenish
        s1.marker.graphicalProperties.line.solidFill = "c9e243"  # Marker outline greenish
        s1.graphicalProperties.line.noFill = True

        xvalues = Reference(ws, min_col=7, min_row=i + 2, max_row=i + 2)
        yvalues = Reference(ws, min_col=8, min_row=i + 2, max_row=i + 2)
        series = Series(values=yvalues, xvalues=xvalues, title="Baseline")
        chart.series.append(series)
        s1 = chart.series[i+1]
        s1.marker.symbol = "diamond"
        s1.marker.size = 10
        s1.marker.graphicalProperties.solidFill = "8187ff"  # Marker filling blue
        s1.marker.graphicalProperties.line.solidFill = "8187ff"  # Marker outline blue
        s1.graphicalProperties.line.noFill = True


    ws.add_chart(chart, "K2")

    return ws

def build_chart_aggregate(wb, td_data_dict):
    ws = wb.active
    chart = ScatterChart()
    chart.title = ' Time deltas overall ' # consider what other titles might be used.
    chart.style = 18
    chart.x_axis.title = 'Time delta for each business case (year intervals)'
    chart.auto_axis = False
    '''this code is necessary to calculate min chart value if its greater than zero'''
    x_axis_min = min_value_all(td_data_dict)
    if x_axis_min >= 0:
        chart.x_axis.scaling.min = 0
    elif x_axis_min < 0:
        anchor = x_axis_min % 365
        chart.x_axis.scaling.min = x_axis_min - anchor
    chart.x_axis.scaling.max = max_value_all(td_data_dict)  # max number (of days) in the x axis. calculated by max_value function
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = len(td_data_dict.keys())  # hard coded for now - although minor issue as number of bc time deltas static
    chart.height = 9  # default is 7.5
    chart.width = 21  # default is 15

    '''changes units on x and y axis'''
    chart.x_axis.majorUnit = 365  # hard coded for now - minor issue as td will normally be in year intervals
    # chart.y_axis.majorUnit = 1.0   testing to see if required

    '''reverses y axis'''
    # chart.x_axis.scaling.orientation = "minMax"
    # chart.y_axis.scaling.orientation = "maxMin"

    '''makes the x axis cross at the max y value'''
    # chart.x_axis.crosses = 'max'

    '''removes lable on y axis'''
    chart.y_axis.delete = True

    # TOD: sort styling
    '''styling chart'''
    '''formating for titles'''
    # font = Font(typeface='Calibri')
    # size = 1200  # 12 point size
    # cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
    # pp = ParagraphProperties(defRPr=cp)
    # rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    # chart.x_axis.title.tx.rich.p[0].pPr = pp  # x_axis title

    # size_2 = 1400
    # cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
    # pp_2 = ParagraphProperties(defRPr=cp_2)
    # rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
    # chart.title.tx.rich.p[0].pPr = pp_2  # chart title

    '''the below assigns series information to the data that has been placed in the chart. 
    old values are placed first show that they show behind the current values'''

    '''there is an unresolved issue here. only thr first 19 series are being placed onto
    the chart. Can not identify the reason'''

    for i in range(0, len(td_data_dict.keys())*2, 2):
        xvalues = Reference(ws, min_col=9, min_row=i + 1, max_row=i + 1)
        yvalues = Reference(ws, min_col=10, min_row=i + 1, max_row=i + 1)
        series = Series(values=yvalues, xvalues=xvalues, title="Current/Actual")
        chart.series.append(series)
        s1 = chart.series[i]
        s1.marker.symbol = "diamond"
        s1.marker.size = 10
        s1.marker.graphicalProperties.solidFill = "c9e243"  # Marker filling greenish
        s1.marker.graphicalProperties.line.solidFill = "c9e243"  # Marker outline greenish
        s1.graphicalProperties.line.noFill = True

        xvalues = Reference(ws, min_col=9, min_row=i + 2, max_row=i + 2)
        yvalues = Reference(ws, min_col=10, min_row=i + 2, max_row=i + 2)
        series = Series(values=yvalues, xvalues=xvalues, title="Baseline")
        chart.series.append(series)
        s1 = chart.series[i + 1]
        s1.marker.symbol = "diamond"
        s1.marker.size = 10
        s1.marker.graphicalProperties.solidFill = "8187ff"  # Marker filling blue
        s1.marker.graphicalProperties.line.solidFill = "8187ff"  # Marker outline blue
        s1.graphicalProperties.line.noFill = True

    ws.add_chart(chart, "K2")

    return wb

def build_chart_master(wb):
    ws = wb.active
    chart = ScatterChart()
    chart.title = ' Time deltas overall ' # consider what other titles might be used.
    chart.style = 18
    chart.x_axis.title = 'Time delta for each business case (year intervals)'
    chart.auto_axis = False
    '''this code is necessary to calculate min chart value if its greater than zero'''
    chart.x_axis.scaling.min = 0     # hard coded
    chart.x_axis.scaling.max = 4000  # hard coded
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 21  # hard coded
    chart.height = 9  # default is 7.5
    chart.width = 21  # default is 15

    '''changes units on x and y axis'''
    chart.x_axis.majorUnit = 365  # hard coded for now - minor issue as td will normally be in year intervals
    # chart.y_axis.majorUnit = 1.0   testing to see if required

    '''reverses y axis'''
    # chart.x_axis.scaling.orientation = "minMax"
    # chart.y_axis.scaling.orientation = "maxMin"

    '''makes the x axis cross at the max y value'''
    # chart.x_axis.crosses = 'max'

    '''removes lable on y axis'''
    chart.y_axis.delete = True

    # TOD: sort styling
    '''styling chart'''
    '''formating for titles'''
    # font = Font(typeface='Calibri')
    # size = 1200  # 12 point size
    # cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
    # pp = ParagraphProperties(defRPr=cp)
    # rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    # chart.x_axis.title.tx.rich.p[0].pPr = pp  # x_axis title

    # size_2 = 1400
    # cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
    # pp_2 = ParagraphProperties(defRPr=cp_2)
    # rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
    # chart.title.tx.rich.p[0].pPr = pp_2  # chart title

    '''the below assigns series information to the data that has been placed in the chart. 
    old values are placed first show that they show behind the current values'''

    row_start = 1
    for i in range(ws.max_row+1):
        for x in range(2):
            ws.cell(row = row_start, column = 5).value = i + 1
            row_start += 1

    for i in range(0, ws.max_row + 1, 2):
        xvalues = Reference(ws, min_col=4, min_row=i + 1, max_row=i + 1)
        yvalues = Reference(ws, min_col=5, min_row=i + 1, max_row=i + 1)
        series = Series(values=yvalues, xvalues=xvalues, title="Current/Actual")
        chart.series.append(series)
        s1 = chart.series[i]
        s1.marker.symbol = "diamond"
        s1.marker.size = 10
        s1.marker.graphicalProperties.solidFill = "c9e243"  # Marker filling greenish
        s1.marker.graphicalProperties.line.solidFill = "c9e243"  # Marker outline greenish
        s1.graphicalProperties.line.noFill = True

        xvalues = Reference(ws, min_col=4, min_row=i + 2, max_row=i + 2)
        yvalues = Reference(ws, min_col=5, min_row=i + 2, max_row=i + 2)
        series = Series(values=yvalues, xvalues=xvalues, title="Baseline")
        chart.series.append(series)
        s1 = chart.series[i + 1]
        s1.marker.symbol = "diamond"
        s1.marker.size = 10
        s1.marker.graphicalProperties.solidFill = "8187ff"  # Marker filling blue
        s1.marker.graphicalProperties.line.solidFill = "8187ff"  # Marker outline blue
        s1.graphicalProperties.line.noFill = True

    ws.add_chart(chart, "K2")

    return wb

'''function for filtering out the group of projects of interest'''
def filter_project_by_group(master_dictionary, value_of_interest):
    output_list = []

    for name in master_dictionary.keys():
        approval_point = master_dictionary[name]['BICC approval point']
        if approval_point == value_of_interest:
            output_list.append(name)

    return output_list

'''function for filter bc/td of interest'''
def filter_bc_for_analysis(any_dict, bc_of_interest):
    output_dict = {}

    for name in any_dict.keys():
        output_list = []
        project_dict = any_dict[name]
        for data in project_dict:
            if data[1][0] == bc_of_interest:
                output_list.append(data)

        output_dict[name] = output_list

    return output_dict

def placing_date_in_excel_plural(milestone_data_dict, td_data_dict):
    wb = Workbook()
    ws = wb.active

    milestone_data_dict # this is left with all data so that everything is printed out into the ws

    '''fairly complicated loop going on here!!'''
    for i, name in enumerate(milestone_data_dict.keys()):
        milestone_master = milestone_data_dict[name]
        milestone_master_length = len(milestone_master)
        project_name_place  = ((i + 1) * milestone_master_length) - (milestone_master_length - 1)
        ws.cell(row= project_name_place, column=1).value = name
        for x in range(milestone_master_length):
            ws.cell(row= x + project_name_place, column=2).value = milestone_master[x][0]
            ws.cell(row= x + project_name_place, column=3).value = milestone_master[x][1][0]
            ws.cell(row= x + project_name_place, column=4).value = milestone_master[x][1][1]

    for i, name in enumerate(td_data_dict.keys()):
        td_master = td_data_dict[name]
        td_master_length = len(td_master)
        project_name_place = ((i + 1) * td_master_length) - (td_master_length - 1)
        ws.cell(row= project_name_place, column=6).value = name
        for x in range (td_master_length):
            ws.cell(row= x + project_name_place, column= 7).value = td_master[x][0]
            ws.cell(row=x + project_name_place, column=8).value = td_master[x][1][0]
            ws.cell(row=x + project_name_place, column=9).value = td_master[x][1][1].days

    '''this section provides an index for the y values part of the chart'''
    project_name_list = list(td_milestone_dict.keys())
    start_row = 1
    for i in range(len(project_name_list)):
        for x in range(2):
            ws.cell(row=start_row, column=10).value = i + 1
            start_row += 1

    return wb

'''function for finding if a milestone is in the past of future. this is very much a hack for now'''
def past_or_future(milestone_dictionary, milestone_of_interest):
    output_list = []

    today = datetime.date.today()

    for name in milestone_dictionary.keys():
        project_list = milestone_dictionary[name]
        for i in range (int(len(project_list)/2)):   # this is half because only want half of the list i.e. the most recent
            if project_list[i][1][0] == milestone_of_interest:
                date_of_interest = project_list[i][1][1]
                try:
                    if today <= date_of_interest:
                        output_list.append(name)
                except TypeError:
                    checking_output = ('No ' + milestone_of_interest + ' data has been provided by ' + str(name))
                    output_list.append(checking_output)

    return output_list

'''function for removing names, its a hack!'''
def remove_names(project_names_list, projects_names_removal_list):
    output_list = []
    for name in project_names_list:
        if name in projects_names_removal_list:
            pass
        else:
            output_list.append(name)

    return output_list

'''function for running individual project td outputs'''
def run(project_name, milestone_data, td_data):
    wb = Workbook()
    ws = wb.active
    print(project_name)
    parser_1 = placing_date_in_excel_single(ws, project_name, milestone_data, td_data)
    approval_point = zero[project_name]['BICC approval point']

    build_chart_single(parser_1, project_name, approval_point, td_data)

    return wb

'''master files are loaded here'''
zero = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx')
one = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_1_2018.xlsx')
two = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_4_2017.xlsx')
three = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2017.xlsx')
four = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2017.xlsx')
five = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_1_2017.xlsx')
six = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_4_2016.xlsx')
last = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2016.xlsx')

'''master files put into a list'''
master_list_all = [zero, one, two, three, four, five, six, last]
master_list_two = [zero, last]

'''FILTERING PROJECTS OF INTEREST. 
This section provides the names of projects that are to be included in output analysis. This is essentially how
different projects are passed into the analysis. For example the programme can return all projects at a particular 
bc stage, or from a particular delivery agency'''
'''all'''
#project_names = zero.keys()
'''bc stage groups via last stage agreed by BICC'''
project_names_1 = filter_project_by_group(zero, 'Strategic Outline Case')
#bc_list = ['Full Business Case', 'pre-Outline Business Case', 'Other', 'Outline Business Case', 'pre-Strategic Outline Case',
# 'Strategic Outline Case'] NOTE: use bc_list function to return all the bcs that are being reported.
'''organisational groups'''
#groups_of_interest = ['Highways England']
'''single - for testing programme'''
# only really required for test
'''removing projects from analysis'''
names_to_remove = ['Commercial Vehicle Services (CVS)', 'Cross Country Rail Franchise Competition',
                   'Crossrail Programme', 'Digital Railway', 'East West Rail Programme (Central Section)',
                   'Future Theory Test Service (FTTS)', 'Great Western Route Modernisation (GWRM) including electrification',
                   'Heathrow Expansion', 'Hexagon', 'High Speed Rail Programme (HS2)',
                   'M20 Lorry Area', 'Mobile Connectivity on Rail Project', 'Network Rail Asset Disposal',
                   'North of England Programme', 'Oxford-Cambridge Expressway', 'Rail Franchising Programme']

project_names = remove_names(project_names_1, names_to_remove)

'''Milestone data that is to be analysed'''
milestone_keys = ['Project MM18', 'Approval MM1', 'Approval MM3', 'Approval MM10', 'Project MM19', 'Project MM20', 'Project MM21']
milestone_dates = ['Project MM18 Forecast - Actual', 'Approval MM1 Forecast / Actual', 'Approval MM3 Forecast / Actual', 'Approval MM10 Forecast / Actual',
                   'Project MM19 Forecast - Actual', 'Project MM20 Forecast - Actual', 'Project MM21 Forecast - Actual']

'''running the programme from here on. steps are are follows:
1) compilation of master data into the python dictionary format. Project names are where the filtering happens. e.g. 
project_names (list) will only contain those projects of interest filtered above.   
2) calculation of time_deltas. based on the milestone_data dictionary that is compiled
3) time_delta milestone dictionaries. This returns a dictionary that contains only the td of interest. e.g. whats the
 td for start of operation or end of project'''

'''step 1'''
milestone_data_1 = milestone_extraction(project_names, master_list_all, milestone_keys, milestone_dates)

'''further filtering'''
checking_project_names_1 = past_or_future(milestone_data_1, 'Start of Operation')
checking_project_names_2 = past_or_future(milestone_data_1, 'Project End Date')
checking_project_names = [x for x in checking_project_names_2 if x not in checking_project_names_1]

#print(checking_project_names)

milestone_data = milestone_extraction(checking_project_names, master_list_all, milestone_keys, milestone_dates)

'''step 2'''
td_data = cal_td(milestone_data_1)
'''step 3'''
td_milestone_dict = filter_bc_for_analysis(td_data, 'FBC - BICC Approval')
#list_of_stages = ['SOBC - BICC Approval', 'OBC - BICC Approval', 'FBC - BICC Approval', 'Start of Construction/build', 'Start of Operation', 'Project End Date']

'''run individual project tds analysis'''
#for project_name in milestone_data.keys():
#    out = run(project_name, milestone_data, td_data)
#    out.save('C:\\Users\\Standalone\\Will\\{}_rcf_test.xlsx'.format(project_name))

'''run aggregate group tds analysis'''
aggregate_wb = placing_date_in_excel_plural(milestone_data_1, td_milestone_dict)
aggregate_chart = build_chart_aggregate(aggregate_wb, td_milestone_dict)

aggregate_chart.save('C:\\Users\\Standalone\\Will\\sobc_to_fbc_final_final.xlsx')

'''master analysis chart compilation'''
#workbook = load_workbook('C:\\Users\\Standalone\\Will\\master_second_next_time_delta.xlsx')
#a = build_chart_master(workbook)
#a.save('C:\\Users\\Standalone\\Will\\master_second_next_time_delta_with_graph.xlsx')